# ============================================================
# HEART FAILURE PREDICTION - EXCEL EXPORT MODULE
# Converts original & cleaned datasets into a formatted .xlsx
# Run: python export_to_excel.py
# Requires: pip install openpyxl pandas
# ============================================================

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
INPUT_CSV    = "heart_failure_clinical_records.csv"
OUTPUT_XLSX  = "heart_failure_datasets.xlsx"
EXCLUDED     = ["diabetes", "smoking", "time"]

# ── Colour palette ────────────────────────────────────────────────────────────
TEAL_DARK   = "1D9E75"
TEAL_LIGHT  = "E1F5EE"
BLUE_DARK   = "185FA5"
BLUE_LIGHT  = "E6F1FB"
EXCL_BG     = "FFF3CD"
EXCL_FG     = "856404"
TARGET_BG   = "FCEBEB"
TARGET_FG   = "A32D2D"
WHITE       = "FFFFFF"
BORDER_COL  = "CCCCCC"

thin   = Side(style="thin", color=BORDER_COL)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COL_WIDTHS = {
    "age": 8, "anaemia": 10, "creatinine_phosphokinase": 22,
    "diabetes": 11, "ejection_fraction": 17, "high_blood_pressure": 18,
    "platelets": 13, "serum_creatinine": 15, "serum_sodium": 13,
    "sex": 7, "smoking": 10, "time": 7, "DEATH_EVENT": 13
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def header_cell(ws, row, col, text, bg, fg="FFFFFF"):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Arial", bold=True, color=fg, size=10)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BORDER

def data_cell(ws, row, col, value, bg=WHITE, fg="000000", bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=9, color=fg, bold=bold)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = BORDER

def set_col_widths(ws, columns):
    for ci, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 14)


# ── Sheet builder ─────────────────────────────────────────────────────────────
def build_data_sheet(ws, df, header_bg, row_bg_even, is_original=False):
    ws.row_dimensions[1].height = 32

    for ci, col in enumerate(df.columns, start=1):
        if col == "DEATH_EVENT":
            header_cell(ws, 1, ci, col, TARGET_BG, TARGET_FG)
        elif is_original and col in EXCLUDED:
            header_cell(ws, 1, ci, col, EXCL_BG, EXCL_FG)
        else:
            header_cell(ws, 1, ci, col, header_bg, WHITE)

    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        alt_bg = row_bg_even if ri % 2 == 0 else WHITE
        for ci, col in enumerate(df.columns, start=1):
            val = row[col]
            if col == "DEATH_EVENT":
                data_cell(ws, ri, ci, int(val), TARGET_BG, TARGET_FG, bold=True)
            elif is_original and col in EXCLUDED:
                data_cell(ws, ri, ci, val, EXCL_BG, EXCL_FG)
            else:
                data_cell(ws, ri, ci, val, alt_bg)

    set_col_widths(ws, df.columns)
    ws.freeze_panes = "A2"


# ── Summary sheet ─────────────────────────────────────────────────────────────
def build_summary_sheet(ws):
    headers = ["Variable", "Data type", "Category", "Included?", "Exclusion reason"]
    for ci, h in enumerate(headers, start=1):
        header_cell(ws, 1, ci, h, TEAL_DARK, WHITE)
    ws.row_dimensions[1].height = 28

    variables = [
        ("age",                      "float64", "Continuous", "Yes", ""),
        ("anaemia",                  "int64",   "Binary",     "Yes", ""),
        ("creatinine_phosphokinase", "int64",   "Continuous", "Yes", ""),
        ("diabetes",                 "int64",   "Binary",     "No",  "Not significant (p=0.4434)"),
        ("ejection_fraction",        "int64",   "Continuous", "Yes", ""),
        ("high_blood_pressure",      "int64",   "Binary",     "Yes", ""),
        ("platelets",                "float64", "Continuous", "Yes", ""),
        ("serum_creatinine",         "float64", "Continuous", "Yes", ""),
        ("serum_sodium",             "int64",   "Continuous", "Yes", ""),
        ("sex",                      "int64",   "Binary",     "Yes", ""),
        ("smoking",                  "int64",   "Binary",     "No",  "Not significant (p=0.5277)"),
        ("time",                     "int64",   "Continuous", "No",  "Data leakage — follow-up period"),
        ("DEATH_EVENT",              "int64",   "Binary",     "Yes", "Prediction target"),
    ]

    for ri, (var, dtype, cat, inc, reason) in enumerate(variables, start=2):
        is_excl   = inc == "No"
        is_target = var == "DEATH_EVENT"
        if is_target:
            rb, rf = TARGET_BG, TARGET_FG
        elif is_excl:
            rb, rf = EXCL_BG, EXCL_FG
        else:
            rb, rf = (TEAL_LIGHT if ri % 2 == 0 else WHITE), "000000"

        for ci, val in enumerate([var, dtype, cat, inc, reason], start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = Font(name="Arial", size=9, color=rf,
                               bold=(ci == 1 or is_target))
            c.fill      = PatternFill("solid", start_color=rb)
            c.alignment = Alignment(horizontal="center" if ci in [3, 4] else "left",
                                    vertical="center")
            c.border    = BORDER

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 36
    ws.freeze_panes = "A2"

    ws.cell(row=16, column=1, value="Colour legend").font = Font(name="Arial", bold=True, size=10)
    for row, label, bg, fg in [
        (17, "Retained variable (original)",  TEAL_DARK,  WHITE),
        (18, "Excluded variable",             EXCL_BG,    EXCL_FG),
        (19, "Target variable (DEATH_EVENT)", TARGET_BG,  TARGET_FG),
        (20, "Retained variable (cleaned)",   BLUE_DARK,  WHITE),
    ]:
        c = ws.cell(row=row, column=1, value=label)
        c.fill      = PatternFill("solid", start_color=bg)
        c.font      = Font(name="Arial", size=9, color=fg)
        c.border    = BORDER
        c.alignment = Alignment(horizontal="left", vertical="center")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"Reading {INPUT_CSV} ...")
    df_original = pd.read_csv(INPUT_CSV)
    df_cleaned  = df_original.drop(columns=EXCLUDED)

    print(f"  Original : {df_original.shape[0]} rows × {df_original.shape[1]} cols")
    print(f"  Cleaned  : {df_cleaned.shape[0]} rows × {df_cleaned.shape[1]} cols")

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Original Dataset (13 cols)"
    build_data_sheet(ws1, df_original, TEAL_DARK, TEAL_LIGHT, is_original=True)

    ws2 = wb.create_sheet("Cleaned Dataset (10 cols)")
    build_data_sheet(ws2, df_cleaned, BLUE_DARK, BLUE_LIGHT, is_original=False)

    ws3 = wb.create_sheet("Summary")
    build_summary_sheet(ws3)

    wb.save(OUTPUT_XLSX)
    print(f"\nSaved → {OUTPUT_XLSX}")
    print("  Sheet 1 : Original Dataset (13 cols)  — teal headers, yellow = excluded")
    print("  Sheet 2 : Cleaned Dataset (10 cols)   — blue headers")
    print("  Sheet 3 : Summary                     — variable overview & colour legend")


if __name__ == "__main__":
    main()